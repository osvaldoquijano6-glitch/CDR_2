"""
core/io.py — Carga unificada de archivos Excel y CSV.

Consolida la lógica de main.py y data/DATA2.PY para leer tablas de datos
sin depender de openpyxl (usa la implementación XML directa como fallback).
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from zipfile import ZipFile
from xml.etree import ElementTree as ET

import pandas as pd


# ─── Namespaces XML ──────────────────────────────────────────────────────────
XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

SUPPORTED_EXTENSIONS = {".xlsx", ".csv"}


# ─── Estructuras de datos ─────────────────────────────────────────────────────
@dataclass(frozen=True)
class PlotColumns:
    time: str
    frequency: str
    active_power: str
    auxiliary_power: str | None = None


# ─── Carga de tabla ───────────────────────────────────────────────────────────
def load_table(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    """Carga un archivo CSV o Excel y devuelve un DataFrame crudo."""
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return _load_csv(path)
    return _load_excel(path, sheet_name)


def _load_csv(path: Path) -> pd.DataFrame:
    sep = _detect_csv_separator(path)
    return pd.read_csv(path, sep=sep, encoding="utf-8-sig")


def _detect_csv_separator(path: Path) -> str:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(4096)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        return dialect.delimiter
    except csv.Error:
        return ";" if sample.count(";") >= sample.count(",") else ","


def _load_excel(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    try:
        return _load_excel_openpyxl(path, sheet_name)
    except ImportError:
        return _load_excel_xml(path, sheet_name)


def _load_excel_openpyxl(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(path, read_only=True, data_only=True)
    target_sheet = _pick_worksheet_name(wb.sheetnames, sheet_name)
    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"La hoja '{ws.title}' no contiene datos.")
    headers = [
        str(v).strip() if v is not None else f"Column {i + 1}"
        for i, v in enumerate(rows[0])
    ]
    return pd.DataFrame([list(r) for r in rows[1:]], columns=headers)


def _load_excel_xml(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    with ZipFile(path) as zf:
        ws_path = _resolve_worksheet_path(zf, sheet_name)
        shared = _read_shared_strings(zf)
        ws = ET.fromstring(zf.read(ws_path))
        sheet_data = ws.find(f"{{{XML_NS}}}sheetData")
        if sheet_data is None:
            raise ValueError(f"Sin datos en '{path.name}'.")

        rows: list[dict[int, object]] = []
        max_col = -1
        for row in sheet_data.findall(f"{{{XML_NS}}}row"):
            row_map: dict[int, object] = {}
            for cell in row.findall(f"{{{XML_NS}}}c"):
                col = _col_index(cell.attrib.get("r", ""))
                row_map[col] = _read_cell(cell, shared)
                max_col = max(max_col, col)
            if row_map:
                rows.append(row_map)

    if not rows:
        raise ValueError(f"Sin filas útiles en '{path.name}'.")

    headers = [
        str(rows[0].get(i, f"Column {i + 1}")).strip() for i in range(max_col + 1)
    ]
    records = [[row.get(i) for i in range(max_col + 1)] for row in rows[1:]]
    return pd.DataFrame(records, columns=headers)


def _resolve_worksheet_path(zf: ZipFile, sheet_name: str | None) -> str:
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        r.attrib["Id"]: r.attrib["Target"]
        for r in rels.findall(f"{{{PKG_REL_NS}}}Relationship")
    }
    sheets = wb.find(f"{{{XML_NS}}}sheets")
    if sheets is None:
        raise ValueError("Sin hojas en el archivo .xlsx.")
    available = []
    sheet_nodes = list(sheets.findall(f"{{{XML_NS}}}sheet"))
    for sheet in sheet_nodes:
        available.append(sheet.attrib.get("name", ""))

    preferred_name = _pick_worksheet_name(available, sheet_name)
    selected = None
    for sheet in sheet_nodes:
        name = sheet.attrib.get("name", "")
        rid = sheet.attrib.get(f"{{{REL_NS}}}id")
        if name == preferred_name:
            selected = rel_map.get(rid)
            break
    if selected is None:
        raise ValueError(f"Hoja '{sheet_name}' no encontrada. Disponibles: {available}")
    return f"xl/{selected.lstrip('/')}"


def _pick_worksheet_name(sheetnames: list[str], requested: str | None) -> str:
    if requested:
        if requested not in sheetnames:
            raise ValueError(f"Hoja '{requested}' no encontrada. Disponibles: {sheetnames}")
        return requested

    if not sheetnames:
        raise ValueError("El archivo Excel no contiene hojas disponibles.")

    preferred = [
        "trend data",
        "low_resolution_5_minutes",
        "low resolution 5 minutes",
        "data",
    ]
    lowered = {name.lower(): name for name in sheetnames}
    for option in preferred:
        if option in lowered:
            return lowered[option]

    ignored_tokens = ("config", "event", "image")
    for name in sheetnames:
        lname = name.lower()
        if not any(token in lname for token in ignored_tokens):
            return name
    return sheetnames[0]


def _read_shared_strings(zf: ZipFile) -> list[str]:
    path = "xl/sharedStrings.xml"
    if path not in zf.namelist():
        return []
    root = ET.fromstring(zf.read(path))
    return [
        "".join(node.text or "" for node in item.iter(f"{{{XML_NS}}}t"))
        for item in root.findall(f"{{{XML_NS}}}si")
    ]


def _col_index(cell_ref: str) -> int:
    letters = "".join(c for c in cell_ref if c.isalpha()).upper()
    idx = 0
    for letter in letters:
        idx = idx * 26 + (ord(letter) - ord("A") + 1)
    return idx - 1


def _read_cell(cell: ET.Element, shared: list[str]) -> object:
    t = cell.attrib.get("t")
    v = cell.find(f"{{{XML_NS}}}v")
    if t == "inlineStr":
        it = cell.find(f"{{{XML_NS}}}is/{{{XML_NS}}}t")
        return it.text if it is not None else None
    if v is None or v.text is None:
        return None
    if t == "s":
        return shared[int(v.text)]
    return v.text


# ─── Detección de columnas ────────────────────────────────────────────────────
def normalize_header(value: object) -> str:
    return " ".join(str(value).strip().lower().split())


def _find_header(headers: list[str], tokens: tuple[str, ...]) -> str | None:
    for h in headers:
        if all(t in normalize_header(h) for t in tokens):
            return h
    return None


def find_matching_header(
    headers: Iterable[str],
    token_groups: Iterable[tuple[str, ...]],
) -> str | None:
    header_list = list(headers)
    for tokens in token_groups:
        match = _find_header(header_list, tokens)
        if match is not None:
            return match
    return None


def detect_time_column(headers: Iterable[str]) -> str:
    header_list = list(headers)
    return (
        _find_header(header_list, ("date", "time"))
        or _find_header(header_list, ("fecha", "hora"))
        or _find_header(header_list, ("time",))
        or header_list[0]
    )


def detect_columns(headers: Iterable[str]) -> PlotColumns:
    headers = list(headers)
    time_col = detect_time_column(headers)
    freq_col = _find_header(headers, ("frequency",)) or _find_header(
        headers, ("frecuencia",)
    )
    power_col = _find_header(headers, ("active", "power")) or _find_header(
        headers, ("potencia", "activa")
    )
    aux_col = (
        _find_header(headers, ("setpoint",))
        or _find_header(headers, ("order",))
        or _find_header(headers, ("theoretical", "power"))
        or _find_header(headers, ("reference", "power"))
        or _find_header(headers, ("potencia", "teor"))
        or _find_header(headers, ("potencia", "refer"))
    )
    missing = []
    if freq_col is None:
        missing.append("frecuencia")
    if missing:
        raise ValueError(
            f"Columnas requeridas no encontradas: {', '.join(missing)}. "
            f"Columnas disponibles: {headers}"
        )
    return PlotColumns(
        time=time_col,
        frequency=freq_col,
        active_power=power_col,
        auxiliary_power=aux_col,
    )


def prepare_signal_dataframe(
    df: pd.DataFrame,
    time_col: str,
    signal_cols: dict[str, str],
) -> pd.DataFrame:
    cleaned = df.copy()
    cleaned.columns = [str(c).strip() for c in cleaned.columns]

    time_col = time_col.strip()
    cleaned[time_col] = _parse_datetime_series(cleaned[time_col])

    keep = [time_col]
    for alias, source_col in signal_cols.items():
        src = source_col.strip()
        cleaned[alias] = pd.to_numeric(cleaned[src], errors="coerce")
        keep.append(alias)

    cleaned = cleaned.dropna(subset=[time_col]).copy()
    value_cols = [name for name in signal_cols]
    cleaned = cleaned.dropna(subset=value_cols, how="all").copy()
    if cleaned.empty:
        raise ValueError(
            "Sin datos válidos después de limpiar la columna de tiempo y señales."
        )

    cleaned = cleaned.sort_values(by=time_col)
    cleaned = cleaned.drop_duplicates(subset=[time_col]).reset_index(drop=True)
    return cleaned[[time_col, *value_cols]].rename(columns={time_col: "time"})


# ─── Preparación del DataFrame ────────────────────────────────────────────────
def prepare_dataframe(df: pd.DataFrame, columns: PlotColumns) -> pd.DataFrame:
    cleaned = df.copy()
    cleaned.columns = [str(c).strip() for c in cleaned.columns]

    time_col = columns.time.strip()
    freq_col = columns.frequency.strip()
    power_col = columns.active_power.strip()
    aux_col = columns.auxiliary_power.strip() if columns.auxiliary_power else None

    cleaned[time_col] = _parse_datetime_series(cleaned[time_col])
    cleaned[freq_col] = pd.to_numeric(cleaned[freq_col], errors="coerce")
    cleaned[power_col] = pd.to_numeric(cleaned[power_col], errors="coerce")
    if aux_col:
        cleaned[aux_col] = pd.to_numeric(cleaned[aux_col], errors="coerce")

    cleaned = cleaned.dropna(subset=[time_col, freq_col, power_col]).copy()
    if cleaned.empty:
        raise ValueError(
            "Sin datos válidos después de limpiar tiempo, frecuencia y potencia."
        )

    cleaned = cleaned.sort_values(by=time_col)
    keep = [time_col, freq_col, power_col]
    if aux_col:
        keep.append(aux_col)
    return cleaned[keep]


def _parse_datetime_series(series: pd.Series) -> pd.Series:
    normalized = series.astype(str).map(_normalize_ampm)
    parsed = pd.Series(pd.NaT, index=normalized.index, dtype="datetime64[ns]")

    ampm_formats = [
        "%d/%m/%Y %I:%M:%S.%f %p",
        "%m/%d/%Y %I:%M:%S.%f %p",
    ]
    for fmt in ampm_formats:
        mask = parsed.isna()
        if not mask.any():
            break
        parsed.loc[mask] = pd.to_datetime(normalized.loc[mask], errors="coerce", format=fmt)

    mask = parsed.isna()
    if mask.any():
        remaining = normalized.loc[mask]
        parsed.loc[mask] = _parse_slash_datetime_candidates(remaining)

    if parsed.isna().any():
        mask = parsed.isna()
        parsed.loc[mask] = pd.to_datetime(normalized.loc[mask], errors="coerce")
    if parsed.isna().all():
        raise ValueError("No se pudo interpretar la columna de tiempo.")
    return parsed


def _parse_slash_datetime_candidates(series: pd.Series) -> pd.Series:
    parsed_dayfirst = _parse_with_formats(
        series,
        [
            "%d/%m/%Y %H:%M:%S:%f",
            "%d/%m/%Y %H:%M:%S.%f",
            "%d/%m/%Y %H:%M:%S",
        ],
    )
    parsed_monthfirst = _parse_with_formats(
        series,
        [
            "%m/%d/%Y %H:%M:%S:%f",
            "%m/%d/%Y %H:%M:%S.%f",
            "%m/%d/%Y %H:%M:%S",
        ],
    )
    return _choose_best_datetime_parse(parsed_dayfirst, parsed_monthfirst)


def _parse_with_formats(series: pd.Series, formats: list[str]) -> pd.Series:
    parsed = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    for fmt in formats:
        mask = parsed.isna()
        if not mask.any():
            break
        parsed.loc[mask] = pd.to_datetime(series.loc[mask], errors="coerce", format=fmt)
    return parsed


def _choose_best_datetime_parse(
    dayfirst: pd.Series,
    monthfirst: pd.Series,
) -> pd.Series:
    day_score = _datetime_parse_score(dayfirst)
    month_score = _datetime_parse_score(monthfirst)
    return dayfirst if day_score >= month_score else monthfirst


def _datetime_parse_score(parsed: pd.Series) -> tuple[int, int, int, float]:
    valid = parsed.dropna()
    if valid.empty:
        return (-1, -1, -1, float("-inf"))

    diffs = valid.diff().dropna()
    negative_jumps = int((diffs < pd.Timedelta(0)).sum())
    if diffs.empty:
        large_jumps = 0
        median_seconds = 0.0
    else:
        positive_diffs = diffs[diffs > pd.Timedelta(0)]
        median_delta = positive_diffs.median() if not positive_diffs.empty else pd.Timedelta(0)
        threshold = max(median_delta * 12, pd.Timedelta(days=2)) if median_delta > pd.Timedelta(0) else pd.Timedelta(days=2)
        large_jumps = int((diffs > threshold).sum())
        median_seconds = float(median_delta.total_seconds()) if median_delta > pd.Timedelta(0) else 0.0

    return (
        int(valid.notna().sum()),
        -negative_jumps,
        -large_jumps,
        -median_seconds,
    )


def _normalize_ampm(value: str) -> str:
    text = str(value).strip()
    for old, new in [
        ("a. m.", "AM"),
        ("p. m.", "PM"),
        ("a.m.", "AM"),
        ("p.m.", "PM"),
        ("a m", "AM"),
        ("p m", "PM"),
    ]:
        text = re.sub(re.escape(old), new, text, flags=re.IGNORECASE)
    return text


def detect_power_unit(column_name: str) -> str:
    import re as _re

    m = _re.search(r"\[(.*?)\]", column_name)
    if m:
        return m.group(1)
    # Heurística: si valores típicos > 100, probablemente kW; si < 100 → MW
    return "kW"
