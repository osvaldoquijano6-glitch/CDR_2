"""Lector de Excel (.xlsx/.xlsm) basado en openpyxl.

Entrega la rejilla cruda de la hoja seleccionada, sin asumir que la primera
fila es encabezado (hay archivos de analizadores con preámbulos de metadatos).
La heurística de selección de hoja proviene del proyecto legado
(core/io.py::_pick_worksheet_name), validada con archivos de campo.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from gcv.ingestion.base import FileFormat, RawDataset, file_sha256

# Hojas típicas de analizadores/registradores usados en campo, en orden de preferencia.
_PREFERRED_SHEETS = [
    "trend data",
    "high_resolution_50_ms",
    "low_resolution_5_minutes",
    "low resolution 5 minutes",
    "data",
]
_IGNORED_TOKENS = ("config", "event", "image")


def pick_sheet(sheetnames: list[str], requested: str | None = None) -> str:
    if requested:
        if requested not in sheetnames:
            raise ValueError(f"Hoja '{requested}' no encontrada. Disponibles: {sheetnames}")
        return requested
    if not sheetnames:
        raise ValueError("El archivo Excel no contiene hojas.")
    lowered = {name.lower(): name for name in sheetnames}
    for option in _PREFERRED_SHEETS:
        if option in lowered:
            return lowered[option]
    for name in sheetnames:
        if not any(token in name.lower() for token in _IGNORED_TOKENS):
            return name
    return sheetnames[0]


def read_excel(path: Path, sheet: str | None = None) -> RawDataset:
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheetnames = list(wb.sheetnames)
        target = pick_sheet(sheetnames, sheet)
        ws = wb[target]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    if not rows:
        raise ValueError(f"La hoja '{target}' de '{path.name}' no contiene datos.")

    grid = pd.DataFrame(rows)
    formato = FileFormat.XLSM if path.suffix.lower() == ".xlsm" else FileFormat.XLSX
    return RawDataset(
        grid=grid,
        source_path=path,
        formato=formato,
        sha256=file_sha256(path),
        hoja=target,
        metadata={"hojas_disponibles": sheetnames},
    )
