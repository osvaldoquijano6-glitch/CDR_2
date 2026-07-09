"""Checklist de Pruebas — réplica exacta del formato REV 1.1 del usuario.

El archivo de salida se abre desde `Checklist_de_Pruebas_REV_1.1.xlsx` y solo
se actualizan las columnas variables por proyecto:

    H  APLICA (SI/NO)
    I  ¿SE PUEDE EJECUTAR LA PRUEBA?
    J  COMENTARIOS REV₁
    K  COMENTARIOS REV₂

Colores, fuentes (Montserrat/Helvetica Neue), celdas combinadas, anchos y
alturas quedan intactos porque nunca se reconstruye la hoja.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from gcv.config.settings import NORMATIVE_DIR

PLANTILLA = NORMATIVE_DIR / "plantillas_usuario" / "Checklist_de_Pruebas_REV_1.1.xlsx"
HOJA = "DATOS POR PRUEBA"

_COL_NUMERO = 4      # D
_COL_APLICA = 8      # H
_COL_EJECUTAR = 9    # I
_COL_COMENT_1 = 10   # J
_COL_COMENT_2 = 11   # K


def universo_pruebas() -> list[dict]:
    """Lee el universo de 45 pruebas desde la plantilla del usuario.

    Devuelve, por prueba: numero, nombre, criterio, tipo (EN SITIO/DOCUMENTAL/NA)
    y los valores por defecto de aplica/ejecutar del formato original.
    """
    wb = openpyxl.load_workbook(PLANTILLA, read_only=True)
    ws = wb[HOJA]
    pruebas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        numero = fila[_COL_NUMERO - 1]
        if numero is None:
            continue
        pruebas.append({
            "numero": int(numero),
            "nombre": str(fila[4] or "").strip(),
            "criterio": str(fila[5] or "").strip(),
            "tipo": str(fila[6] or "").strip(),
            "aplica": str(fila[_COL_APLICA - 1] or "").strip(),
            "ejecutar": str(fila[_COL_EJECUTAR - 1] or "").strip(),
        })
    wb.close()
    return pruebas


def generar_checklist(
    destino: Path,
    aplica: dict[int, str] | None = None,
    ejecutar: dict[int, str] | None = None,
    comentarios_rev1: dict[int, str] | None = None,
    comentarios_rev2: dict[int, str] | None = None,
) -> Path:
    """Genera el checklist con el formato exacto del usuario.

    Cada dict mapea número de prueba (1–45) → texto de la celda. Las pruebas
    no incluidas conservan el valor de la plantilla.
    """
    wb = openpyxl.load_workbook(PLANTILLA)
    ws = wb[HOJA]

    actualizaciones = {
        _COL_APLICA: aplica or {},
        _COL_EJECUTAR: ejecutar or {},
        _COL_COMENT_1: comentarios_rev1 or {},
        _COL_COMENT_2: comentarios_rev2 or {},
    }
    for fila in ws.iter_rows(min_row=2):
        valor_num = fila[_COL_NUMERO - 1].value
        if valor_num is None:
            continue
        numero = int(valor_num)
        for col, datos in actualizaciones.items():
            if numero in datos:
                ws.cell(row=fila[0].row, column=col, value=datos[numero])

    destino = Path(destino)
    if destino.suffix.lower() != ".xlsx":
        destino.mkdir(parents=True, exist_ok=True)
        destino = destino / "Checklist_de_Pruebas.xlsx"
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    return destino
